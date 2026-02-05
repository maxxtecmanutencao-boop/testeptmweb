import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
import io
import time
import shutil
import os


st.set_page_config(
	page_title='PTM JSL - Sistema de Consultas',
	layout='wide',
	initial_sidebar_state='expanded',
	menu_items={
		'About': 'Sistema de Consultas PTM JSL - Versão 2.0'
	}
)


def get_theme_css(theme_name="Original"):
	"""Retorna o CSS customizado baseado no tema selecionado"""
	
	themes = {
		"Original": {
			"bg_gradient": "linear-gradient(135deg, #050506 0%, #09090b 50%, #0000 100%)",
			"primary_color": "#00a8ff",
			"secondary_color": "#00ff88",
			"accent_color": "#ff6b6b",
			"text_color": "#f7fbff",
			"card_bg": "linear-gradient(135deg, #0e0e10 0%, #1a1a1a 100%)",
		},
		"Azul Escuro / Verde Neon": {
			"bg_gradient": "linear-gradient(135deg, #001a33 0%, #003366 50%, #004d99 100%)",
			"primary_color": "#00ffff",
			"secondary_color": "#00ff00",
			"accent_color": "#00ff88",
			"text_color": "#e0ffff",
			"card_bg": "linear-gradient(135deg, #001f3f 0%, #003d5c 100%)",
		},
		"Vermelho Escuro / Amarelo": {
			"bg_gradient": "linear-gradient(135deg, #330000 0%, #660000 50%, #990000 100%)",
			"primary_color": "#ffff00",
			"secondary_color": "#ffcc00",
			"accent_color": "#ff6600",
			"text_color": "#ffffcc",
			"card_bg": "linear-gradient(135deg, #4d0000 0%, #800000 100%)",
		},
		"Neon Futurista": {
			"bg_gradient": "linear-gradient(135deg, #0a0015 0%, #1a0033 50%, #2d004d 100%)",
			"primary_color": "#ff00ff",
			"secondary_color": "#00ffff",
			"accent_color": "#ffff00",
			"text_color": "#ffffff",
			"card_bg": "linear-gradient(135deg, #1a0033 0%, #330066 100%)",
		}
	}
	
	theme = themes.get(theme_name, themes["Original"])
	
	return f"""
<style>
	@import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700;900&family=Roboto:wght@300;400;700&display=swap');

	.main {{
		background: {theme["bg_gradient"]};
		color: {theme["text_color"]};
	}}

	.stApp {{
		background: {theme["bg_gradient"]};
	}}

	h1, h2, h3 {{
		font-family: 'Orbitron', sans-serif !important;
		color: {theme["text_color"]} !important;
		text-shadow: 0 0 6px rgba(0,0,0,0.6);
		letter-spacing: 2px;
	}}

	.metric-card {{
		background: {theme["card_bg"]};
		border-radius: 15px;
		padding: 20px;
		box-shadow: 0 10px 40px rgba(0,0,0,0.7);
		border: 2px solid rgba(255,255,255,0.05);
		transition: all 0.2s ease;
		margin: 10px 0;
	}}

	.metric-card:hover {{
		transform: translateY(-5px);
		box-shadow: 0 16px 56px rgba(0,0,0,0.8);
		border-color: rgba(255,255,255,0.08);
	}}

	.metric-value {{
		font-size: 48px;
		font-weight: 900;
		font-family: 'Orbitron', sans-serif;
		color: {theme["secondary_color"]};
		text-shadow: none;
	}}

	.metric-label {{
		font-size: 16px;
		color: {theme["text_color"]};
		text-transform: uppercase;
		letter-spacing: 1px;
	}}

	.stButton>button {{
		background: linear-gradient(135deg, {theme["primary_color"]} 0%, {theme["primary_color"]}dd 100%);
		color: white;
		border: none;
		border-radius: 25px;
		padding: 12px 30px;
		font-weight: bold;
		font-family: 'Orbitron', sans-serif;
		text-transform: uppercase;
		letter-spacing: 1px;
		box-shadow: 0 6px 18px rgba(0,0,0,0.6);
		transition: all 0.2s ease;
	}}

	.stButton>button:hover {{
		background: linear-gradient(135deg, {theme["secondary_color"]} 0%, {theme["secondary_color"]}dd 100%);
		box-shadow: 0 8px 28px rgba(0,0,0,0.75);
		transform: translateY(-2px);
	}}

	.stDownloadButton>button {{
		background: linear-gradient(135deg, {theme["accent_color"]} 0%, {theme["accent_color"]}dd 100%);
		color: white;
		border: none;
		border-radius: 25px;
		padding: 12px 30px;
		font-weight: bold;
		font-family: 'Orbitron', sans-serif;
		text-transform: uppercase;
		letter-spacing: 1px;
		box-shadow: 0 4px 15px rgba(255, 107, 107, 0.3);
		transition: all 0.2s ease;
	}}

	.stDownloadButton>button:hover {{
		background: linear-gradient(135deg, {theme["accent_color"]}dd 0%, {theme["accent_color"]}aa 100%);
		box-shadow: 0 6px 25px rgba(255, 107, 107, 0.6);
		transform: translateY(-2px);
	}}

	.dataframe {{
		border-radius: 10px;
		overflow: hidden;
		box-shadow: 0 6px 28px rgba(0,0,0,0.65);
	}}

	.status-badge {{
		display: inline-block;
		padding: 5px 15px;
		border-radius: 20px;
		font-weight: bold;
		font-size: 12px;
		text-transform: uppercase;
		letter-spacing: 1px;
	}}

	.alert-box {{
		background: linear-gradient(135deg, #ff6b6b 0%, #ee5a6f 100%);
		border-left: 5px solid #ff0000;
		padding: 15px;
		border-radius: 10px;
		margin: 10px 0;
		box-shadow: 0 6px 22px rgba(255, 107, 107, 0.25);
	}}

	.success-box {{
		background: linear-gradient(135deg, #00ff88 0%, #00cc66 100%);
		border-left: 5px solid #00ff00;
		padding: 15px;
		border-radius: 10px;
		margin: 10px 0;
		box-shadow: 0 6px 22px rgba(0, 255, 136, 0.25);
	}}

	.sidebar .sidebar-content {{
		background: linear-gradient(180deg, #09090b 0%, #0000 100%);
	}}

	.stRadio > label {{
		font-family: 'Orbitron', sans-serif;
		color: {theme["primary_color"]} !important;
		font-weight: bold;
	}}

	.stTextInput>div>div>input, .stTextArea>div>div>textarea {{
		background-color: rgba(255,255,255,0.03);
		border: 1px solid rgba(255,255,255,0.06);
		border-radius: 10px;
		color: {theme["text_color"]};
		font-family: 'Roboto', sans-serif;
	}}

	.stTextInput>div>div>input:focus, .stTextArea>div>div>textarea:focus {{
		border-color: {theme["primary_color"]};
		box-shadow: 0 0 12px rgba(0,168,255,0.25);
	}}

	.header-container {{
		background: {theme["card_bg"]};
		padding: 20px;
		border-radius: 15px;
		box-shadow: 0 12px 40px rgba(0,0,0,0.6);
		margin-bottom: 30px;
		border: 2px solid rgba(255,255,255,0.06);
	}}

	.pulse {{
		animation: pulse 2s infinite;
	}}

	@keyframes pulse {{
		0% {{ opacity: 1; }}
		50% {{ opacity: 0.6; }}
		100% {{ opacity: 1; }}
	}}

	.glow {{
		animation: glow 2s ease-in-out infinite alternate;
	}}

	@keyframes glow {{
		from {{ text-shadow: 0 0 5px rgba(0,168,255,0.6); }}
		to {{ text-shadow: 0 0 10px rgba(0,168,255,0.65); }}
	}}
	
	.clock-display {{
		font-family: 'Orbitron', sans-serif;
		font-size: 28px;
		color: {theme["secondary_color"]};
		text-shadow: none;
		font-weight: bold;
		text-align: center;
		padding: 10px;
		background: rgba(255,255,255,0.02);
		border-radius: 10px;
		border: 1px solid rgba(255,255,255,0.04);
	}}

	footer {{visibility: hidden;}}
	.viewerBadge_container__1QSob {{display: none;}}
	#MainMenu {{visibility: hidden;}}
	footer:after {{
		content:''; 
		visibility: visible;
		display: block;
	}}
</style>
"""


# FUNÇÕES AUXILIARES (após os imports)

def load_bd(path: Path, sheet_name='BDBI'):
	"""Carrega dados da planilha BD.xlsm (ou BD.xlsx) de uma aba específica SEM CACHE"""
	try:
		df_bd = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
	except Exception:
		try:
			df_bd = pd.read_excel(path, sheet_name=sheet_name)
		except:
			df_bd = pd.read_excel(path, engine="openpyxl")
	return df_bd


def load_bd_no_cache(path: Path, sheet_name='BDBI'):
	"""Carrega dados sem cache para operações de atualização"""
	try:
		df_bd = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
	except Exception:
		try:
			df_bd = pd.read_excel(path, sheet_name=sheet_name)
		except:
			df_bd = pd.read_excel(path, engine="openpyxl")
	return df_bd


def normalize_columns(df: pd.DataFrame):
	mapping = {}
	cols = {c.lower(): c for c in df.columns}

	def find(key_parts):
		for part in key_parts:
			for c_low, c in cols.items():
				if part in c_low:
					return c
		return None

	mapping_candidates = {
		'Pedido': ['pedido', 'order', 'ordem'],
		'Item': ['item', 'posicao'],
		'Data do Pedido': ['data do pedido', 'data_pedido', 'pedido data', 'order date', 'data_ped'],
		'Data Prevista': ['entrega'],
		'Entrega Real': ['entrega real'],
		'Situação': ['situa', 'status'],
		'Remessa': ['remessa', 'shipment'],
		'Origem Ativo': ['origem', 'origem ativo'],
		'Destino Ativo': ['destino', 'destino ativo'],
		'NFe': ['nfe', 'nf-e', 'nota fiscal'],
		'DTM': ['dtm'],
		'Fase Atual': ['fase atual', 'fase_atual', 'fase']
	}

	for target, parts in mapping_candidates.items():
		found = find(parts)
		if found:
			mapping[found] = target

	if mapping:
		df = df.rename(columns=mapping)
	return df


def parse_dates(df: pd.DataFrame):
	for col in ['Data do Pedido', 'Data Prevista', 'Entrega Real']:
		if col in df.columns:
			df[col] = pd.to_datetime(df[col], errors='coerce')
	return df


def status_contains(val, keyword):
	if pd.isna(val):
		return False
	return keyword.lower() in str(val).lower()


def highlight_alerts(row, alert_statuses, days_threshold=7):
	try:
		status = str(row.get('Situação', ''))
		if any(status_contains(status, s) for s in alert_statuses):
			dped = row.get('Data do Pedido')
			dprev = row.get('Data Prevista')
			if pd.notna(dped) and pd.notna(dprev):
				delta = (dprev - dped).days
				if delta <= days_threshold:
					return ['background-color: rgba(255, 107, 107, 0.3); border-left: 3px solid #ff0000'] * len(row)
	except Exception:
		pass
	return [''] * len(row)


# ---- PTMs em atrasos (funções auxiliares) ----

def carregar_dados_remessas(arquivo='BD.xlsm'):
	"""Carrega e processa os dados do Excel para análise de remessas"""
	try:
		df = pd.read_excel(arquivo)
		
		df = df[df['Pedido'].notna()].copy()
		
		colunas_data = ['Data do Pedido', 'Data Prevista', 'Entrega Real']
		for col in colunas_data:
			if col in df.columns:
				df[col] = pd.to_datetime(df[col], errors='coerce')
		
		return df
	except Exception as e:
		st.error(f"Erro ao carregar arquivo: {e}")
		return pd.DataFrame()


def calcular_metricas_remessas(df):
	"""Calcula métricas e status das remessas"""
	if df.empty:
		return df
	
	df = df.copy()
	hoje = pd.Timestamp(datetime.now().date())
	
	df['Dias desde Pedido'] = (hoje - df['Data do Pedido']).dt.days
	df['Dias até Entrega'] = (df['Data Prevista'] - hoje).dt.days
	df['Prazo Total (dias)'] = (df['Data Prevista'] - df['Data do Pedido']).dt.days
	
	def determinar_status(row):
		if pd.notna(row['Entrega Real']):
			if row['Entrega Real'] <= row['Data Prevista']:
				return '✅ Entregue no Prazo'
			else:
				dias_atraso = (row['Entrega Real'] - row['Data Prevista']).days
				return f'⚠️ Entregue com Atraso ({dias_atraso}d)'
		else:
			dias_restantes = row['Dias até Entrega']
			if dias_restantes < 0:
				return f'🔴 PRAZO VENCIDO ({abs(dias_restantes)}d)'
			elif dias_restantes <= 7:
				return f'🟡 ATENÇÃO - Próximo ao Prazo ({dias_restantes}d)'
			else:
				return f'🟢 No Prazo ({dias_restantes}d)'
	
	df['Status'] = df.apply(determinar_status, axis=1)
	
	if 'Fase Atual' in df.columns:
		df['Tempo Parado (dias)'] = df.apply(
			lambda row: row['Dias desde Pedido'] if pd.notna(row['Fase Atual']) else 0,
			axis=1
		)
	else:
		df['Tempo Parado (dias)'] = 0
	
	return df

def gerar_estatisticas_remessas(df):
	"""Gera estatísticas resumidas"""
	if df is None or df.empty:
		return {}

	stats = {
		'total': len(df),
		'entregues': len(df[df['Status'].str.contains('Entregue', na=False)]),
		'atrasados': len(df[df['Status'].str.contains('VENCIDO|PRAZO VENCIDO', na=False)]),
		'atencao': len(df[df['Status'].str.contains('ATENÇÃO', na=False)]),
		'no_prazo': len(df[df['Status'].str.contains('No Prazo', na=False)]),
		'tempo_medio_parado': df['Tempo Parado (dias)'].mean() if 'Tempo Parado (dias)' in df.columns else 0
	}
	return stats


def formatar_data_abreviada(data_col):
	"""Formata datas para formato abreviado dd/mm/aa"""
	if pd.isna(data_col):
		return ""
	try:
		return pd.to_datetime(data_col).strftime('%d/%m/%y')
	except:
		return ""


def create_metric_card(label, value, icon="📊"):
	return f"""
	<div class="metric-card">
		<div style="display: flex; align-items: center; justify-content: space-between;">
			<div>
				<div class="metric-label">{icon} {label}</div>
				<div class="metric-value">{value}</div>
			</div>
		</div>
	</div>
	"""


def limpar_backups_antigos(base_path, max_backups=3):
	"""Mantém apenas os últimos 3 backups mais recentes"""
	try:
		backup_files = list(base_path.glob('BD_backup_*.xlsm')) + list(base_path.glob('BD_backup_*.xlsx'))
		
		if len(backup_files) > max_backups:
			backup_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
			
			for backup_file in backup_files[max_backups:]:
				backup_file.unlink()
				st.info(f"🗑️ Backup antigo removido: {backup_file.name}")
	except Exception as e:
		st.warning(f"⚠️ Aviso ao limpar backups antigos: {e}")


def sincronizar_ptm26_para_bdbi(bd_path: Path):
	"""
	Sincroniza dados da aba PTM26 para a aba BDBI
	
	REGRAS ESPECIAIS DE FASE ATUAL:
	1. Todos os registros da PTM26 têm Fase Atual considerada como [8] Em Transporte
	2. Na BDBI, o campo "Fase Atual" é atualizado para [8] Em Transporte
	3. EXCETO quando o registro na BDBI já estiver [10] Finalizado - neste caso mantém Finalizado
	4. Outros campos são atualizados normalmente se houver diferença
	"""
	try:
		from openpyxl import load_workbook
		
		# Carrega ambas as abas
		df_bdbi = pd.read_excel(bd_path, sheet_name='BDBI', engine='openpyxl')
		df_ptm26 = pd.read_excel(bd_path, sheet_name='PTM26', engine='openpyxl')
		
		# Normaliza as colunas
		df_bdbi = normalize_columns(df_bdbi)
		df_ptm26 = normalize_columns(df_ptm26)
		
		# Verifica se as colunas chave existem
		if 'Pedido' not in df_bdbi.columns or 'Item' not in df_bdbi.columns:
			return False, "❌ Colunas 'Pedido' ou 'Item' não encontradas na aba BDBI"
		
		if 'Pedido' not in df_ptm26.columns or 'Item' not in df_ptm26.columns:
			return False, "❌ Colunas 'Pedido' ou 'Item' não encontradas na aba PTM26"
		
		# Identifica a coluna "Fase Atual" em ambas as abas
		fase_atual_col_bdbi = None
		fase_atual_col_ptm26 = None
		
		for col in df_bdbi.columns:
			if 'fase' in col.lower() and 'atual' in col.lower():
				fase_atual_col_bdbi = col
				break
		
		for col in df_ptm26.columns:
			if 'fase' in col.lower() and 'atual' in col.lower():
				fase_atual_col_ptm26 = col
				break
		
		if not fase_atual_col_bdbi:
			st.warning("⚠️ Coluna 'Fase Atual' não encontrada na BDBI. A sincronização continuará sem atualizar este campo.")
		
		# Cria chave de merge
		df_bdbi['_merge_key'] = df_bdbi['Pedido'].astype(str) + '_' + df_bdbi['Item'].astype(str)
		df_ptm26['_merge_key'] = df_ptm26['Pedido'].astype(str) + '_' + df_ptm26['Item'].astype(str)
		
		# Contadores
		registros_atualizados = 0
		registros_mantidos_finalizados = 0
		registros_para_em_transporte = 0
		campos_atualizados = []
		
		# Para cada registro da PTM26
		for idx_ptm, row_ptm in df_ptm26.iterrows():
			merge_key = row_ptm['_merge_key']
			
			# Busca o registro correspondente na BDBI
			mask_bdbi = df_bdbi['_merge_key'] == merge_key
			
			if mask_bdbi.any():
				idx_bdbi = df_bdbi[mask_bdbi].index[0]
				campos_modificados = []
				
				# LÓGICA ESPECIAL PARA FASE ATUAL
				if fase_atual_col_bdbi:
					fase_atual_bdbi = str(df_bdbi.loc[idx_bdbi, fase_atual_col_bdbi])
					
					# Verifica se está [10] Finalizado
					if '[10]' in fase_atual_bdbi.upper() or 'FINALIZADO' in fase_atual_bdbi.upper():
						# Mantém [10] Finalizado
						registros_mantidos_finalizados += 1
					else:
						# Atualiza para [8] Em Transporte
						valor_anterior = df_bdbi.loc[idx_bdbi, fase_atual_col_bdbi]
						df_bdbi.loc[idx_bdbi, fase_atual_col_bdbi] = '[8] Em Transporte'
						if str(valor_anterior) != '[8] Em Transporte':
							campos_modificados.append(fase_atual_col_bdbi)
						registros_para_em_transporte += 1
				
				# Atualiza outros campos normalmente
				for col in df_ptm26.columns:
					if col == '_merge_key':
						continue
					
					# Pula a coluna Fase Atual da PTM26 (já tratada acima)
					if col == fase_atual_col_ptm26:
						continue
					
					if col in df_bdbi.columns:
						valor_ptm = row_ptm[col]
						valor_bdbi = df_bdbi.loc[idx_bdbi, col]
						
						# Compara valores (tratando NaN)
						if pd.isna(valor_ptm) and pd.isna(valor_bdbi):
							continue
						elif pd.isna(valor_ptm) or pd.isna(valor_bdbi):
							df_bdbi.loc[idx_bdbi, col] = valor_ptm
							campos_modificados.append(col)
						elif str(valor_ptm) != str(valor_bdbi):
							df_bdbi.loc[idx_bdbi, col] = valor_ptm
							campos_modificados.append(col)
				
				if campos_modificados:
					registros_atualizados += 1
					campos_atualizados.extend(campos_modificados)
		
		# Remove a coluna auxiliar
		df_bdbi = df_bdbi.drop(columns=['_merge_key'])
		
		# Salva o arquivo atualizado preservando outras abas e o formato XLSM
		from openpyxl.utils.dataframe import dataframe_to_rows
		
		book = load_workbook(bd_path)
		
		# Remove a aba BDBI se existir
		if 'BDBI' in book.sheetnames:
			std = book['BDBI']
			book.remove(std)
		
		# Cria nova aba BDBI com os dados atualizados
		ws = book.create_sheet('BDBI', 0)
		
		# Escreve os dados da aba BDBI
		for r_idx, row in enumerate(dataframe_to_rows(df_bdbi, index=False, header=True), 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
		
		# Salva preservando o formato XLSM
		book.save(bd_path)
		book.close()
		
		campos_unicos = list(set(campos_atualizados))
		
		mensagem_resultado = f"""✅ **Sincronização concluída com sucesso!**

📊 **Estatísticas:**
- **{registros_atualizados}** registros atualizados na BDBI
- **{registros_para_em_transporte}** registros alterados para [8] Em Transporte
- **{registros_mantidos_finalizados}** registros mantidos como [10] Finalizado
- **Campos modificados:** {', '.join(campos_unicos) if campos_unicos else 'Nenhum'}

🔄 **Regras aplicadas:**
✓ Registros da PTM26 → [8] Em Transporte na BDBI
✓ Registros já [10] Finalizado → Mantidos como Finalizado
"""
		
		return True, mensagem_resultado
	
	except Exception as e:
		return False, f"❌ Erro na sincronização: {str(e)}"


def main():
	# Inicializa o tema no session_state
	if 'theme' not in st.session_state:
		st.session_state.theme = "Original"
	
	# Aplica o CSS do tema selecionado
	st.markdown(get_theme_css(st.session_state.theme), unsafe_allow_html=True)
	
	base_path = Path(__file__).resolve().parent
	img1 = base_path / 'Petrobras.png'
	img2 = base_path / 'logo jsl.png'

	st.markdown('<div class="header-container">', unsafe_allow_html=True)
	cols_hdr = st.columns([1, 6, 1])
	with cols_hdr[0]:
		if img1.exists():
			st.image(str(img1), width=140)
	with cols_hdr[1]:
		st.markdown('<h1 style="text-align: center;" class="glow">🚀 SISTEMA PTM JSL 2.0</h1>', unsafe_allow_html=True)
		st.markdown('<p style="text-align: center; color: #b0c4de; font-size: 14px;">Sistema Avançado de Consultas e Monitoramento</p>', unsafe_allow_html=True)
	with cols_hdr[2]:
		if img2.exists():
			st.image(str(img2), width=140)
	st.markdown('</div>', unsafe_allow_html=True)

	bd_path = base_path / 'BD.xlsm'

	# Verifica também por BD.xlsx para compatibilidade
	if not bd_path.exists():
		bd_path_xlsx = base_path / 'BD.xlsx'
		if bd_path_xlsx.exists():
			bd_path = bd_path_xlsx

	# MODIFICAÇÃO CRÍTICA: Sempre carrega dados atualizados da BDBI (sem cache)
	if bd_path.exists():
		df = load_bd(bd_path, sheet_name='BDBI')
	else:
		st.warning('⚠️ Arquivo BD.xlsm não encontrado. Faça upload do arquivo.')
		uploaded = st.file_uploader('📁 Upload BD.xlsm', type=['xlsm', 'xlsx'])
		if uploaded is None:
			st.stop()
		df = pd.read_excel(uploaded, sheet_name='BDBI', engine='openpyxl')

	df = normalize_columns(df)
	df = parse_dates(df)

	show_cols = ['Pedido', 'Item', 'Data do Pedido', 'Data Prevista', 'Entrega Real', 'Situação', 'Remessa', 'Origem Ativo', 'Destino Ativo', 'NFe', 'DTM', 'Fase Atual']
	available_cols = [c for c in show_cols if c in df.columns]

	def format_for_display(df_in: pd.DataFrame):
		df_out = df_in.copy()
		for col in ['Data do Pedido', 'Entrega Real']:
			if col in df_out.columns:
				try:
					df_out[col] = pd.to_datetime(df_out[col], errors='coerce').dt.strftime('%d/%m/%Y')
				except Exception:
					pass
		return df_out

	def prepare_export(df_in: pd.DataFrame):
		df_out = df_in.copy()
		for col in ['Data do Pedido', 'Entrega Real']:
			if col in df_out.columns:
				try:
					df_out[col] = pd.to_datetime(df_out[col], errors='coerce').dt.date
				except Exception:
					pass
		export_cols = [c for c in available_cols if c in df_out.columns]
		return df_out[export_cols]

	st.sidebar.markdown('## 🎛️ PAINEL DE CONTROLE')
	# Garantir key única por sessão para evitar StreamlitDuplicateElementKey em execuções paralelas
	if 'nav_radio_key' not in st.session_state:
		import uuid
		st.session_state['nav_radio_key'] = f"nav_radio_{uuid.uuid4().hex[:8]}"
	page = st.sidebar.radio('📍 Navegação', [
		'🏠 Dashboard', 
		'📊 Resumo por Status', 
		'📋 Planilha Completa', 
		'📈 Analytics', 
		'🔄 Atualizar Sistema BD',
		'🔄 Sincronizar PTM26 → BDBI',
		'PTMs em atrasos'
	], label_visibility='visible', key=st.session_state['nav_radio_key'])

	status_list = ['Aguard. DSM', 'Aguard. NFe', 'Aguard. Coleta', 'Aguard. Remessa', 'Finalizado', 'Em Transporte', 'Aguard. MIGO']
	status_keys = ['aguard. dsm', 'aguard. nfe', 'aguard. coleta', 'aguard. remessa', 'finalizado', 'em transporte', 'aguard. migo']
	alert_statuses = ['aguard. dsm', 'aguard. nfe', 'aguard. coleta']

	if page == '🏠 Dashboard':
		# Seletor de Tema
		st.markdown("### 🎨 Selecione o Tema")
		theme_options = ["Original", "Azul Escuro / Verde Neon", "Vermelho Escuro / Amarelo", "Neon Futurista"]
		selected_theme = st.selectbox("Escolha o tema visual:", theme_options, index=theme_options.index(st.session_state.theme), key="theme_selector")
		
		if selected_theme != st.session_state.theme:
			st.session_state.theme = selected_theme
			st.rerun()
		
		st.markdown("---")
		
		# Data + Relógio
		current_datetime = datetime.now()
		date_str = current_datetime.strftime('%d/%m/%Y')
		time_str = current_datetime.strftime('%H:%M:%S')
		
		st.markdown(f'''
		<div class="clock-display">
			📅 {date_str}   |   🕐 {time_str}
		</div>
		''', unsafe_allow_html=True)
		
		st.markdown("<br>", unsafe_allow_html=True)

		total_records = len(df)

		if 'Fase Atual' in df.columns:
			status_counts = {}
			for label, key in zip(status_list, status_keys):
				count = df['Fase Atual'].astype(str).str.contains(key, case=False, na=False).sum()
				status_counts[label] = count
		else:
			status_counts = {label: 0 for label in status_list}

		col1, col2, col3, col4 = st.columns(4)

		with col1:
			st.markdown(create_metric_card("Total de Registros", total_records, "📦"), unsafe_allow_html=True)

		with col2:
			finalizados = status_counts.get('Finalizado', 0)
			st.markdown(create_metric_card("Finalizados", finalizados, "✅"), unsafe_allow_html=True)

		with col3:
			em_transporte = status_counts.get('Em Transporte', 0)
			st.markdown(create_metric_card("Em Transporte", em_transporte, "🚚"), unsafe_allow_html=True)

		with col4:
			aguard_nfe = status_counts.get('Aguard. NFe', 0)
			st.markdown(create_metric_card("Aguardando NFe", aguard_nfe, "📄"), unsafe_allow_html=True)

		st.markdown("---")

		col5, col6 = st.columns(2)

		with col5:
			aguard_dsm = status_counts.get('Aguard. DSM', 0)
			st.markdown(create_metric_card("Aguardando DSM", aguard_dsm, "📋"), unsafe_allow_html=True)

		with col6:
			aguard_coleta = status_counts.get('Aguard. Coleta', 0)
			st.markdown(create_metric_card("Aguardando Coleta", aguard_coleta, "📦"), unsafe_allow_html=True)

		st.markdown("---")

		col_chart1, col_chart2 = st.columns(2)

		with col_chart1:
			st.markdown("### 📊 Distribuição por Status")
			if 'Situação' in df.columns:
				counts = df['Situação'].astype(str).value_counts().reset_index()
				counts.columns = ['Situação', 'Quantidade']
				fig = px.pie(counts, values='Quantidade', names='Situação',
						hole=0.4,
						color_discrete_sequence=px.colors.sequential.Plasma)
				fig.update_layout(
					paper_bgcolor='rgba(0,0,0,0)',
					plot_bgcolor='rgba(0,0,0,0)',
					font=dict(color='white', family='Roboto')
				)
				st.plotly_chart(fig, use_container_width=True)

		with col_chart2:
			st.markdown("### 📈 Status em Barra")
			status_df = pd.DataFrame(list(status_counts.items()), columns=['Status', 'Quantidade'])
			fig = px.bar(status_df, x='Status', y='Quantidade',
					color='Quantidade',
					color_continuous_scale='Turbo',
					text='Quantidade')
			fig.update_layout(
				paper_bgcolor='rgba(0,0,0,0)',
				plot_bgcolor='rgba(0,0,0,0)',
				font=dict(color='white', family='Roboto'),
				xaxis=dict(tickangle=-45)
			)
			st.plotly_chart(fig, use_container_width=True)

		st.markdown("### ⚠️ Alertas Críticos")
		alert_count = 0
		for alert_status in alert_statuses:
			mask = df['Situação'].astype(str).str.contains(alert_status, case=False, na=False) if 'Situação' in df.columns else pd.Series([False] * len(df))
			df_alert = df[mask]
			for _, row in df_alert.iterrows():
				try:
					dped = row.get('Data do Pedido')
					dprev = row.get('Data Prevista')
					if pd.notna(dped) and pd.notna(dprev):
						delta = (dprev - dped).days
						if delta <= 7:
							alert_count += 1
				except:
					pass

		if alert_count > 0:
			st.markdown(f'<div class="alert-box">🚨 <strong>{alert_count}</strong> pedidos com prazo crítico (≤ 7 dias)</div>', unsafe_allow_html=True)
		else:
			st.markdown('<div class="success-box">✅ Nenhum alerta crítico no momento</div>', unsafe_allow_html=True)

	elif page == '📊 Resumo por Status':
		st.markdown('## 🔍 BUSCA E CONSULTA')

		st.markdown('### 🔎 Busca por Remessa')
		if 'remessa_query' not in st.session_state:
			st.session_state.remessa_query = ''

		modo = st.radio('🎯 Modo de busca', ['✍️ Manual (texto)', '📋 Selecionar da lista'], horizontal=True, key='modo_busca')

		remessa_choices = []
		if 'Remessa' in df.columns:
			remessa_choices = sorted(df['Remessa'].dropna().astype(str).unique())

		cols1, cols2 = st.columns([3, 1])
		with cols1:
			if modo == '📋 Selecionar da lista' and remessa_choices:
				sel = st.multiselect('Selecione uma ou mais Remessas', remessa_choices, default=[])
				manual_text = ''
			else:
				manual_text = st.text_area('Digite remessas (separadas por vírgula ou uma por linha)', value=st.session_state.remessa_query, height=100)
				sel = []

		with cols2:
			st.markdown("<br>", unsafe_allow_html=True)
			if st.button('🔍 Buscar', use_container_width=True):
				if sel:
					st.session_state.remessa_query = ','.join(sel)
				else:
					st.session_state.remessa_query = manual_text or ''
			if st.button('🗑️ Limpar', use_container_width=True):
				st.session_state.remessa_query = ''
				sel = []
				manual_text = ''
				
		if st.session_state.remessa_query:
			raw = st.session_state.remessa_query
			keys = [k.strip() for k in raw.replace('\n', ',').split(',') if k.strip()]
			mask = pd.Series(False, index=df.index)
			if 'Remessa' in df.columns and keys:
				for k in keys:
					mask = mask | df['Remessa'].astype(str).str.contains(k, case=False, na=False)
			df_search = df[mask]

			st.markdown(f'### 📋 Resultados: <span class="metric-value" style="font-size: 32px;">{len(df_search)}</span> linhas', unsafe_allow_html=True)

			if not df_search.empty:
				display_df = format_for_display(df_search[available_cols])
				st.dataframe(display_df, use_container_width=True, height=400)

				try:
					towrite = io.BytesIO()
					with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
						df_export = prepare_export(df_search)
						df_export.to_excel(writer, index=False, sheet_name='Consultas')
					towrite.seek(0)
					st.download_button(
						label='📥 Exportar Excel',
						data=towrite.getvalue(),
						file_name=f'consulta_remessas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
						mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
						use_container_width=True
					)
				except Exception:
					st.error('❌ Erro ao preparar arquivo Excel para download.')
			else:
				st.info('ℹ️ Nenhum resultado encontrado.')
		else:
			st.markdown('### 📊 Visualização por Status')
			for label, key in zip(status_list, status_keys):
				with st.expander(f"📌 {label}", expanded=False):
					mask = df['Situação'].astype(str).str.contains(key, case=False, na=False) if 'Situação' in df.columns else pd.Series([False] * len(df))
					df_status = df[mask]
					if df_status.empty:
						st.write('ℹ️ Nenhum registro')
					else:
						st.markdown(f'**Total: {len(df_status)} registros | Status: {label}**')

						display_cols = ['Situação'] + [c for c in available_cols if c != 'Situação']
						display_cols = [c for c in display_cols if c in df_status.columns]

						sty = df_status[display_cols].style.apply(lambda r: highlight_alerts(r, alert_statuses), axis=1)
						fmt = {}
						for _c in ['Data do Pedido', 'Entrega Real']:
							if _c in df_status.columns:
								fmt[_c] = lambda v: v.strftime('%d/%m/%Y') if pd.notna(v) else ''
						if fmt:
							sty = sty.format(fmt)
						st.dataframe(sty, use_container_width=True, height=300)

	elif page == '📋 Planilha Completa':
		st.markdown('## 📋 PLANILHA COMPLETA')

		col_filter1, col_filter2 = st.columns(2)

		with col_filter1:
			if 'Situação' in df.columns:
				situacoes = ['Todos'] + sorted(df['Situação'].dropna().astype(str).unique().tolist())
				situacao_filter = st.selectbox('🔽 Filtrar por Situação', situacoes)
			else:
				situacao_filter = 'Todos'

		with col_filter2:
			if 'Origem Ativo' in df.columns:
				origens = ['Todos'] + sorted(df['Origem Ativo'].dropna().astype(str).unique().tolist())
				origem_filter = st.selectbox('🔽 Filtrar por Origem', origens)
			else:
				origem_filter = 'Todos'

		df_filtered = df.copy()
		df_filtered_display = df_filtered.copy()

		if situacao_filter != 'Todos':
			df_filtered = df_filtered[df_filtered['Situação'].astype(str) == situacao_filter]
			df_filtered_display = df_filtered_display[df_filtered_display['Situação'].astype(str) == situacao_filter]

		if origem_filter != 'Todos':
			df_filtered = df_filtered[df_filtered['Origem Ativo'].astype(str) == origem_filter]
			df_filtered_display = df_filtered_display[df_filtered_display['Origem Ativo'].astype(str) == origem_filter]

		st.markdown(f'### 📊 Exibindo {len(df_filtered_display)} de {len(df)} registros')

		st.markdown('#### ✏️ Editar Dados (Clique duas vezes na célula para editar)')

		edited_df = st.data_editor(
			df_filtered_display,
			use_container_width=True,
			height=500,
			num_rows="fixed",
			disabled=[col for col in df_filtered_display.columns if col != 'Situação'],
			column_config={
				"Situação": st.column_config.SelectboxColumn(
					"Situação",
					help="Clique para alterar a situação",
					options=sorted(df['Situação'].dropna().astype(str).unique().tolist()),
					required=True,
				)
			}
		)

		if st.button('💾 Salvar Alterações no Banco de Dados', type='primary', use_container_width=True):
			try:
				for idx, row in edited_df.iterrows():
					original_idx = df_filtered_display.index[idx]
					df.loc[original_idx, 'Situação'] = row['Situação']

				# Salva na aba BDBI preservando outras abas
				from openpyxl import load_workbook
				
				book = load_workbook(bd_path)
				if 'BDBI' in book.sheetnames:
					std = book['BDBI']
					book.remove(std)
				book.save(bd_path)
				book.close()
				
				with pd.ExcelWriter(bd_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
					df.to_excel(writer, sheet_name='BDBI', index=False)

				st.success('✅ Alterações salvas com sucesso na aba BDBI!')
				st.balloons()
				time.sleep(1)
				st.rerun()
			except Exception as e:
				st.error(f'❌ Erro ao salvar alterações: {str(e)}')

		try:
			towrite = io.BytesIO()
			with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
				edited_df.to_excel(writer, index=False, sheet_name='Dados')
			towrite.seek(0)
			st.download_button(
				label='📥 Exportar Planilha Filtrada',
				data=towrite.getvalue(),
				file_name=f'planilha_completa_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
				mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
				use_container_width=True
			)
		except Exception:
			st.error('❌ Erro ao preparar arquivo Excel.')

	elif page == '📈 Analytics':
		st.markdown('## 📈 ANALYTICS AVANÇADO')

		tem_dados_atraso = ('Data Prevista' in df.columns and 'Situação' in df.columns)
		
		if tem_dados_atraso:
			tab1, tab2, tab3 = st.tabs(["📊 Análise de Status", "📅 Timeline de Entregas", "⏰ Análise de Atrasos"])
		else:
			tab1, tab2 = st.tabs(["📊 Análise de Status", "📅 Timeline de Entregas"])

		with tab1:
			st.markdown('### 📊 Distribuição por Status em Tempo Real')

			if 'Situação' in df.columns:
				counts = df['Situação'].astype(str).value_counts().reset_index()
				counts.columns = ['Situação', 'Quantidade']
				counts = counts.sort_values('Quantidade', ascending=False)

				col_metrics = st.columns(len(counts))
				for idx, (_, row) in enumerate(counts.iterrows()):
					with col_metrics[idx % len(col_metrics)]:
						st.markdown(create_metric_card(row['Situação'], row['Quantidade'], "📊"), unsafe_allow_html=True)

				st.markdown("---")

				col_g1, col_g2 = st.columns(2)

				with col_g1:
					fig = px.bar(counts, x='Situação', y='Quantidade',
							text='Quantidade',
							color='Quantidade',
							color_continuous_scale=[[0, '#00ff88'], [0.5, '#00d9ff'], [1, '#ff00ff']])
					fig.update_traces(
						texttemplate='%{text}', 
						textposition='outside',
						marker=dict(
							line=dict(color='#00d9ff', width=2)
						)
					)
					fig.update_layout(
						paper_bgcolor='rgba(0,0,0,0)',
						plot_bgcolor='rgba(15,15,30,0.8)',
						font=dict(color='#00ff88', family='Orbitron', size=14),
						xaxis=dict(
							tickangle=-45, 
							title='Status',
							gridcolor='rgba(0, 217, 255, 0.2)',
							showgrid=True
						),
						yaxis=dict(
							title='Quantidade',
							gridcolor='rgba(0, 217, 255, 0.2)',
							showgrid=True
						),
						showlegend=False,
						height=500,
						title=dict(
							text='Gráfico de Barras Digital',
							font=dict(color='#00d9ff', size=18)
						)
					)
					st.plotly_chart(fig, use_container_width=True)

				with col_g2:
					colors_neon = ['#00ff88', '#00d9ff', '#ff00ff', '#ffff00', '#ff0080', '#00ffff', '#ff6600']
					fig = px.pie(counts, values='Quantidade', names='Situação',
							hole=0.5,
							color_discrete_sequence=colors_neon)
					fig.update_traces(
						textposition='inside', 
						textinfo='percent+label+value',
						marker=dict(line=dict(color='#00d9ff', width=3))
					)
					fig.update_layout(
						paper_bgcolor='rgba(0,0,0,0)',
						plot_bgcolor='rgba(0,0,0,0)',
						font=dict(color='#00ff88', family='Orbitron', size=14),
						showlegend=True,
						legend=dict(
							font=dict(color='white'),
							bgcolor='rgba(26, 26, 46, 0.8)',
							bordercolor='#00d9ff',
							borderwidth=2
						),
						height=500,
						title=dict(
							text='Gráfico de Rosca Digital',
							font=dict(color='#00d9ff', size=18),
							x=0.5
						)
					)
					st.plotly_chart(fig, use_container_width=True)

		with tab2:
			st.markdown('### 📅 Timeline de Entregas - Evolução Mensal')

			if 'Entrega Real' in df.columns:
				df_timeline = df[df['Entrega Real'].notna()].copy()

				if not df_timeline.empty:
					df_timeline['Mes'] = pd.to_datetime(df_timeline['Entrega Real']).dt.to_period('M').astype(str)
					timeline_counts = df_timeline.groupby('Mes').size().reset_index(name='Quantidade')

					col_timeline = st.columns([1, 3, 1])
					with col_timeline[1]:
						st.markdown(create_metric_card("Total de Entregas", len(df_timeline), "📦"), unsafe_allow_html=True)

					st.markdown("---")

					fig = px.line(timeline_counts, x='Mes', y='Quantidade',
							markers=True,
							line_shape='spline')
					fig.update_traces(
						line=dict(color='#00d9ff', width=3),
						marker=dict(size=12, color='#00ff88', line=dict(width=2, color='white'))
					)
					fig.update_layout(
						paper_bgcolor='rgba(0,0,0,0)',
						plot_bgcolor='rgba(0,0,0,0)',
						font=dict(color='white', family='Roboto', size=12),
						xaxis=dict(tickangle=-45, title='Mês'),
						yaxis=dict(title='Quantidade de Entregas'),
						hovermode='x unified',
						height=500
					)
					st.plotly_chart(fig, use_container_width=True)

					st.markdown("---")
					st.markdown("### 📊 Estatísticas de Entregas")

					col_stats = st.columns(4)
					with col_stats[0]:
						media = timeline_counts['Quantidade'].mean()
						st.markdown(create_metric_card("Média Mensal", f"{media:.0f}", "📊"), unsafe_allow_html=True)
					with col_stats[1]:
						maximo = timeline_counts['Quantidade'].max()
						st.markdown(create_metric_card("Máximo Mensal", f"{maximo:.0f}", "📈"), unsafe_allow_html=True)
					with col_stats[2]:
						minimo = timeline_counts['Quantidade'].min()
						st.markdown(create_metric_card("Mínimo Mensal", f"{minimo:.0f}", "📉"), unsafe_allow_html=True)
					with col_stats[3]:
						if 'Remessa' in df_timeline.columns:
							remessas = df_timeline['Remessa'].nunique()
							st.markdown(create_metric_card("Remessas Únicas", remessas, "🚚"), unsafe_allow_html=True)
				else:
					st.info('ℹ️ Nenhuma entrega registrada.')
			else:
				st.info('ℹ️ Coluna "Entrega Real" não encontrada.')

		if tem_dados_atraso:
			with tab3:
				st.markdown('### ⏰ ANÁLISE DETALHADA DE ATRASOS')
				
				try:
					df_atraso = df.copy()
					hoje = pd.Timestamp.now()
					
					df_atraso = df_atraso[df_atraso['Data Prevista'].notna()].copy()
					
					if len(df_atraso) == 0:
						st.warning("⚠️ Não há dados com Data Prevista válida para análise de atrasos.")
					else:
						df_atraso['Dias_Atraso'] = (hoje - df_atraso['Data Prevista']).dt.days
						
						mask_atrasado = (df_atraso['Data Prevista'] < hoje) & (~df_atraso['Situação'].astype(str).str.contains('finalizado', case=False, na=False))
						df_atrasados = df_atraso[mask_atrasado].copy()
						
						if 'Entrega Real' in df.columns:
							df_atraso_com_entrega = df_atraso[df_atraso['Entrega Real'].notna()].copy()
							if len(df_atraso_com_entrega) > 0:
								mask_finalizado_atrasado = (df_atraso_com_entrega['Situação'].astype(str).str.contains('finalizado', case=False, na=False)) & \
															   (df_atraso_com_entrega['Entrega Real'] > df_atraso_com_entrega['Data Prevista'])
								df_finalizados_atrasados = df_atraso_com_entrega[mask_finalizado_atrasado].copy()
								if len(df_finalizados_atrasados) > 0:
									df_finalizados_atrasados['Dias_Atraso'] = (df_finalizados_atrasados['Entrega Real'] - df_finalizados_atrasados['Data Prevista']).dt.days
							else:
								df_finalizados_atrasados = pd.DataFrame()
						else:
							df_finalizados_atrasados = pd.DataFrame()
						
						st.markdown("#### 📊 Visão Geral de Atrasos")
						
						col_m1, col_m2, col_m3, col_m4 = st.columns(4)
						
						with col_m1:
							total_atrasados_aberto = len(df_atrasados)
							st.markdown(create_metric_card("Atrasados (Em Aberto)", total_atrasados_aberto, "🔴"), unsafe_allow_html=True)
						
						with col_m2:
							total_finalizados_atraso = len(df_finalizados_atrasados)
							st.markdown(create_metric_card("Finalizados com Atraso", total_finalizados_atraso, "🟠"), unsafe_allow_html=True)
						
						with col_m3:
							if len(df_atrasados) > 0:
								media_atraso = df_atrasados['Dias_Atraso'].mean()
								st.markdown(create_metric_card("Média de Atraso", f"{media_atraso:.0f} dias", "📅"), unsafe_allow_html=True)
							else:
								st.markdown(create_metric_card("Média de Atraso", "0 dias", "📅"), unsafe_allow_html=True)
						
						with col_m4:
							if len(df_atrasados) > 0:
								max_atraso = df_atrasados['Dias_Atraso'].max()
								st.markdown(create_metric_card("Maior Atraso", f"{max_atraso:.0f} dias", "⚠️"), unsafe_allow_html=True)
							else:
								st.markdown(create_metric_card("Maior Atraso", "0 dias", "⚠️"), unsafe_allow_html=True)
						
						st.markdown("---")
						
						st.markdown("### 🔴 Pedidos Atrasados em Aberto (Por Status Atual)")
						
						if len(df_atrasados) > 0:
							atraso_por_status = df_atrasados.groupby('Situação').agg({
								'Pedido': 'count',
								'Dias_Atraso': ['mean', 'max', 'min']
							}).round(1)
							
							atraso_por_status.columns = ['Quantidade', 'Média Dias Atraso', 'Máx Dias Atraso', 'Mín Dias Atraso']
							atraso_por_status = atraso_por_status.sort_values('Quantidade', ascending=False).reset_index()
							
							st.markdown("#### 📌 Detalhamento por Situação Atual")
							
							for idx, row in atraso_por_status.iterrows():
								with st.expander(f"🔴 {row['Situação']} - {int(row['Quantidade'])} pedidos atrasados", expanded=(idx == 0)):
									col_det1, col_det2, col_det3, col_det4 = st.columns(4)
									
									with col_det1:
										st.metric("📦 Quantidade", int(row['Quantidade']))
									with col_det2:
										st.metric("📊 Média de Atraso", f"{row['Média Dias Atraso']:.0f} dias")
									with col_det3:
										st.metric("📈 Maior Atraso", f"{row['Máx Dias Atraso']:.0f} dias")
									with col_det4:
										st.metric("📉 Menor Atraso", f"{row['Mín Dias Atraso']:.0f} dias")
									
									mask_status = df_atrasados['Situação'] == row['Situação']
									cols_to_show = ['Pedido', 'Item', 'Data Prevista', 'Dias_Atraso', 'Remessa', 'Origem Ativo', 'Destino Ativo']
									cols_available = [c for c in cols_to_show if c in df_atrasados.columns]
									df_status_detalhe = df_atrasados[mask_status][cols_available].copy()
									df_status_detalhe = df_status_detalhe.sort_values('Dias_Atraso', ascending=False)
									if 'Data Prevista' in df_status_detalhe.columns:
										df_status_detalhe['Data Prevista'] = pd.to_datetime(df_status_detalhe['Data Prevista']).dt.strftime('%d/%m/%Y')
									
									st.dataframe(df_status_detalhe, use_container_width=True, height=250)
							
							st.markdown("---")
							
							st.markdown("### 📊 Visualização de Atrasos")
							
							fig = px.bar(atraso_por_status, 
										x='Situação', 
										y='Quantidade',
										text='Quantidade',
										color='Média Dias Atraso',
										color_continuous_scale=[[0, '#00ff88'], [0.5, '#ffff00'], [1, '#ff0080']])
							fig.update_traces(texttemplate='%{text}', textposition='outside')
							fig.update_layout(
								paper_bgcolor='rgba(0,0,0,0)',
								plot_bgcolor='rgba(15,15,30,0.8)',
								font=dict(color='#00ff88', family='Orbitron', size=12),
								xaxis=dict(tickangle=-45, title='Status Atual', gridcolor='rgba(0, 217, 255, 0.2)'),
								yaxis=dict(title='Quantidade de Pedidos Atrasados', gridcolor='rgba(0, 217, 255, 0.2)'),
								height=500
							)
							st.plotly_chart(fig, use_container_width=True)
							
							st.markdown("---")
							
							fig2 = px.bar(atraso_por_status, 
										 x='Situação', 
										 y='Média Dias Atraso',
										 text='Média Dias Atraso',
										 color='Média Dias Atraso',
										 color_continuous_scale=[[0, '#00ff88'], [0.5, '#ffff00'], [1, '#ff0080']])
							fig2.update_traces(texttemplate='%{text:.1f}', textposition='outside')
							fig2.update_layout(
								paper_bgcolor='rgba(0,0,0,0)',
								plot_bgcolor='rgba(15,15,30,0.8)',
								font=dict(color='#00ff88', family='Orbitron', size=12),
								xaxis=dict(tickangle=-45, title='Status Atual', gridcolor='rgba(0, 217, 255, 0.2)'),
								yaxis=dict(title='Média de Dias de Atraso', gridcolor='rgba(0, 217, 255, 0.2)'),
								height=500
							)
							st.plotly_chart(fig2, use_container_width=True)
							
						else:
							st.success("✅ Excelente! Não há pedidos atrasados em aberto no momento.")
						
						st.markdown("---")
						
						st.markdown("### 🟠 Pedidos Finalizados com Atraso (Histórico)")
						
						if len(df_finalizados_atrasados) > 0:
							col_fin1, col_fin2, col_fin3 = st.columns(3)
							
							with col_fin1:
								st.markdown(create_metric_card("Total Finalizados c/ Atraso", len(df_finalizados_atrasados), "🟠"), unsafe_allow_html=True)
							
							with col_fin2:
								media_atraso_fin = df_finalizados_atrasados['Dias_Atraso'].mean()
								st.markdown(create_metric_card("Média de Atraso", f"{media_atraso_fin:.0f} dias", "📊"), unsafe_allow_html=True)
							
							with col_fin3:
								max_atraso_fin = df_finalizados_atrasados['Dias_Atraso'].max()
								st.markdown(create_metric_card("Maior Atraso", f"{max_atraso_fin:.0f} dias", "📈"), unsafe_allow_html=True)
							
							st.markdown("---")
							
							cols_to_show_fin = ['Pedido', 'Item', 'Data Prevista', 'Entrega Real', 'Dias_Atraso', 'Remessa', 'Origem Ativo']
							cols_available_fin = [c for c in cols_to_show_fin if c in df_finalizados_atrasados.columns]
							df_fin_display = df_finalizados_atrasados[cols_available_fin].copy()
							df_fin_display = df_fin_display.sort_values('Dias_Atraso', ascending=False).head(20)
							if 'Data Prevista' in df_fin_display.columns:
								df_fin_display['Data Prevista'] = pd.to_datetime(df_fin_display['Data Prevista']).dt.strftime('%d/%m/%Y')
							if 'Entrega Real' in df_fin_display.columns:
								df_fin_display['Entrega Real'] = pd.to_datetime(df_fin_display['Entrega Real']).dt.strftime('%d/%m/%Y')
							
							st.markdown("#### 📋 Top 20 Pedidos Finalizados com Maior Atraso")
							st.dataframe(df_fin_display, use_container_width=True, height=400)
							
						else:
							st.success("✅ Não há registro de pedidos finalizados com atraso.")
						
						st.markdown("---")
						
						st.markdown("### 📊 Resumo Comparativo")
						
						col_comp1, col_comp2 = st.columns(2)
						
						with col_comp1:
							st.markdown("#### 🔴 Situação Crítica - Em Aberto")
							if len(df_atrasados) > 0:
								cols_criticos = ['Pedido', 'Situação', 'Dias_Atraso', 'Remessa']
								cols_criticos_avail = [c for c in cols_criticos if c in df_atrasados.columns]
								df_criticos = df_atrasados.nlargest(5, 'Dias_Atraso')[cols_criticos_avail].copy()
								for idx_crit, row in df_criticos.iterrows():
									st.markdown(f"""
									<div style='background: rgba(255, 107, 107, 0.2); padding: 10px; border-left: 4px solid #ff0000; margin: 5px 0; border-radius: 5px;'>
										<strong>🔴 Pedido {row['Pedido']}</strong><br>
										<small>Status: {row.get('Situação', 'N/A')} | Atraso: <strong>{row['Dias_Atraso']:.0f} dias</strong></small><br>
										<small>Remessa: {row.get('Remessa', 'N/A')}</small>
									</div>
									""", unsafe_allow_html=True)
							else:
								st.success("✅ Sem pedidos críticos")
						
						with col_comp2:
							st.markdown("#### 🟢 Pedidos no Prazo")
							mask_no_prazo = (df_atraso['Data Prevista'] >= hoje) | (df_atraso['Situação'].astype(str).str.contains('finalizado', case=False, na=False))
							total_no_prazo = len(df_atraso[mask_no_prazo])
							percentual = (total_no_prazo / len(df_atraso) * 100) if len(df_atraso) > 0 else 0
							
							st.markdown(create_metric_card("Total no Prazo", total_no_prazo, "🟢"), unsafe_allow_html=True)
							st.markdown(create_metric_card("Percentual", f"{percentual:.1f}%", "📊"), unsafe_allow_html=True)
				
				except Exception as e:
					st.error(f"❌ Erro ao processar análise de atrasos: {str(e)}")
					st.info("ℹ️ Verifique se os dados estão no formato correto (datas válidas).")

	elif page == '🔄 Atualizar Sistema BD':
		st.markdown('## 🔄 ATUALIZAÇÃO DO SISTEMA BD')
		st.markdown('Atualize a base de dados do sistema com a planilha mais recente')
		st.markdown("---")
		
		st.subheader("📊 Informações da Base Atual")
		
		col_info1, col_info2, col_info3 = st.columns(3)
		
		with col_info1:
			st.markdown(create_metric_card("Total de Registros", len(df), "📦"), unsafe_allow_html=True)
		
		with col_info2:
			if 'Data do Pedido' in df.columns:
				ultima_atualizacao = df['Data do Pedido'].max()
				if pd.notna(ultima_atualizacao):
					data_formatada = ultima_atualizacao.strftime('%d/%m/%Y')
				else:
					data_formatada = "N/A"
			else:
				data_formatada = "N/A"
			st.markdown(create_metric_card("Última Atualização", data_formatada, "📅"), unsafe_allow_html=True)
		
		with col_info3:
			if bd_path.exists():
				tamanho_mb = os.path.getsize(bd_path) / (1024 * 1024)
				st.markdown(create_metric_card("Tamanho do Arquivo", f"{tamanho_mb:.2f} MB", "💾"), unsafe_allow_html=True)
			else:
				st.markdown(create_metric_card("Tamanho do Arquivo", "N/A", "💾"), unsafe_allow_html=True)
		
		st.markdown("---")
		
		st.subheader("📋 Upload Manual")
		
		st.info("""
		**Passo a passo para atualização manual:**
		
		1. Baixe a planilha BD.xlsm mais recente
		2. Faça o upload do arquivo usando o botão abaixo
		3. Confirme a substituição da base de dados
		
		⚠️ **Atenção:** A operação de upload irá substituir completamente a base de dados atual!
		""")
		
		st.markdown("### 📤 Upload do Arquivo BD.xlsm")
		
		uploaded_file = st.file_uploader(
			"Selecione o arquivo BD.xlsm",
			type=['xlsm', 'xlsx'],
			help="Faça upload apenas de arquivos Excel (.xlsm ou .xlsx)",
			key="upload_bd_manual"
		)
		
		if uploaded_file is not None:
			st.markdown("### 👁️ Preview dos Dados")
			
			try:
				df_novo = pd.read_excel(uploaded_file, sheet_name='BDBI', engine='openpyxl')
				df_novo = normalize_columns(df_novo)
				df_novo = parse_dates(df_novo)
				
				col_preview1, col_preview2, col_preview3 = st.columns(3)
				
				with col_preview1:
					st.markdown(create_metric_card("Novos Registros", len(df_novo), "📊"), unsafe_allow_html=True)
				
				with col_preview2:
					colunas_count = len(df_novo.columns)
					st.markdown(create_metric_card("Colunas", colunas_count, "📑"), unsafe_allow_html=True)
				
				with col_preview3:
					diferenca = len(df_novo) - len(df)
					icon = "⬆️" if diferenca > 0 else "⬇️" if diferenca < 0 else "➡️"
					st.markdown(create_metric_card("Diferença", f"{diferenca:+d}", icon), unsafe_allow_html=True)
				
				st.markdown("---")
				
				st.markdown("#### 📋 Primeiras 10 linhas do arquivo:")
				display_cols_preview = available_cols if all(col in df_novo.columns for col in available_cols) else list(df_novo.columns[:10])
				display_df_preview = format_for_display(df_novo.head(10)[display_cols_preview])
				st.dataframe(display_df_preview, use_container_width=True)
				
				st.markdown("---")
				
				col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
				
				with col_btn2:
					confirmar_upload = st.button(
						'✅ CONFIRMAR E SUBSTITUIR BASE DE DADOS',
						type='primary',
						use_container_width=True,
						key="confirmar_upload_manual"
					)
				
				if confirmar_upload:
					try:
						backup_path = base_path / f'BD_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
						
						if bd_path.exists():
							shutil.copy2(bd_path, backup_path)
							st.success(f"✅ Backup criado: {backup_path.name}")
							
							limpar_backups_antigos(base_path, max_backups=3)
						
						# Salva o arquivo completo (todas as abas)
						with open(bd_path, 'wb') as f:
							f.write(uploaded_file.getbuffer())
						
						st.success("✅ Base de dados atualizada com sucesso!")
						st.balloons()
						
						st.info("🔄 Recarregando o sistema com os novos dados...")
						time.sleep(2)
						st.rerun()
						
					except Exception as e:
						st.error(f"❌ Erro ao atualizar a base de dados: {str(e)}")
						
			except Exception as e:
				st.error(f"❌ Erro ao ler o arquivo: {str(e)}")
				st.warning("⚠️ Verifique se o arquivo está no formato correto (.xlsm ou .xlsx)")
		
		else:
			st.warning("⚠️ Nenhum arquivo selecionado. Faça o upload do arquivo BD.xlsm para continuar.")
		
		st.markdown("---")
		
		st.subheader("💾 Backups Disponíveis")
		
		backup_files = list(base_path.glob('BD_backup_*.xlsm')) + list(base_path.glob('BD_backup_*.xlsx'))
		
		if backup_files:
			backup_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
			
			st.markdown(f"**Total de backups encontrados:** {len(backup_files)}")
			st.markdown("*Mostrando os 5 backups mais recentes. Apenas os 3 últimos são mantidos automaticamente.*")
			
			for backup_file in backup_files[:5]:
				col_backup1, col_backup2, col_backup3 = st.columns([2, 1, 1])
				
				with col_backup1:
					file_size = os.path.getsize(backup_file) / (1024 * 1024)
					file_date = datetime.fromtimestamp(backup_file.stat().st_mtime).strftime('%d/%m/%Y %H:%M')
					st.text(f"📁 {backup_file.name}")
					st.caption(f"📅 {file_date} | 💾 {file_size:.2f} MB")
				
				with col_backup2:
					if st.button(f"🔄 Restaurar", key=f"restore_{backup_file.name}", use_container_width=True):
						try:
							shutil.copy2(backup_file, bd_path)
							st.success(f"✅ Base restaurada do backup: {backup_file.name}")
							time.sleep(2)
							st.rerun()
						except Exception as e:
							st.error(f"❌ Erro ao restaurar backup: {str(e)}")
				
				with col_backup3:
					if st.button(f"🗑️ Excluir", key=f"delete_{backup_file.name}", use_container_width=True):
						try:
							backup_file.unlink()
							st.success(f"✅ Backup excluído: {backup_file.name}")
							time.sleep(1)
							st.rerun()
						except Exception as e:
							st.error(f"❌ Erro ao excluir backup: {str(e)}")
		else:
			st.info("ℹ️ Nenhum backup disponível no momento.")

	elif page == '🔄 Sincronizar PTM26 → BDBI':
		st.markdown('## 🔄 SINCRONIZAÇÃO PTM26 → BDBI')
		st.markdown('Sincronize os dados atualizados manualmente da aba PTM26 para a aba principal BDBI')
		st.markdown("---")
		
		st.info("""
		### 📋 Como funciona a sincronização:
		
		1. **Aba PTM26**: Aba de trabalho onde você atualiza os dados manualmente durante a semana
		2. **Aba BDBI**: Aba principal que alimenta todo o dashboard
		3. **Sincronização**: Copia os dados da PTM26 para a BDBI com regras especiais
		
		### 🔄 REGRAS ESPECIAIS DE "FASE ATUAL":
		
		✅ **Todos os registros da PTM26** são considerados como **[8] Em Transporte**
		
		✅ **Na BDBI, a "Fase Atual" será atualizada para [8] Em Transporte**
		
		⚠️ **EXCETO**: Registros que já estão **[10] Finalizado** na BDBI serão **mantidos como Finalizado**
		
		📝 **Outros campos** são atualizados normalmente conforme os valores da PTM26
		
		💾 **Um backup automático será criado antes da sincronização**
		""")
		
		st.markdown("---")
		
		# Informações das abas
		try:
			df_bdbi_info = load_bd_no_cache(bd_path, sheet_name='BDBI')
			df_ptm26_info = load_bd_no_cache(bd_path, sheet_name='PTM26')
			
			col_aba1, col_aba2 = st.columns(2)
			
			with col_aba1:
				st.markdown("### 📊 Aba BDBI (Principal)")
				st.markdown(create_metric_card("Registros", len(df_bdbi_info), "📦"), unsafe_allow_html=True)
				
				# Verifica quantos estão Finalizados
				fase_col = None
				for col in df_bdbi_info.columns:
					if 'fase' in col.lower() and 'atual' in col.lower():
						fase_col = col
						break
				
				if fase_col:
					finalizados_count = df_bdbi_info[fase_col].astype(str).str.contains('[10]|finalizado', case=False, na=False).sum()
					st.info(f"🔒 **{finalizados_count}** registros com [10] Finalizado (serão mantidos)")
				
				st.success("✅ Aba BDBI encontrada")
			
			with col_aba2:
				st.markdown("### 📝 Aba PTM26 (Manual)")
				st.markdown(create_metric_card("Registros", len(df_ptm26_info), "📋"), unsafe_allow_html=True)
				st.info(f"🚚 Todos os {len(df_ptm26_info)} registros → [8] Em Transporte")
				st.success("✅ Aba PTM26 encontrada")
			
			st.markdown("---")
			
			# Exemplo visual
			st.markdown("### 📖 Exemplo de Sincronização:")
			exemplo_data = {
				'Situação': ['Caso 1', 'Caso 2', 'Caso 3'],
				'BDBI (Antes)': ['[5] Qualquer fase', '[8] Em Transporte', '[10] Finalizado ✅'],
				'PTM26 (Manual)': ['Dados atualizados', 'Dados atualizados', 'Dados atualizados'],
				'BDBI (Depois)': ['[8] Em Transporte 🔄', '[8] Em Transporte ✓', '[10] Finalizado 🔒']
			}
			df_exemplo = pd.DataFrame(exemplo_data)
			st.dataframe(df_exemplo, use_container_width=True, hide_index=True)
			
			st.markdown("---")
			
			# Botão de sincronização
			col_sync1, col_sync2, col_sync3 = st.columns([1, 2, 1])
			
			with col_sync2:
				if st.button("🔄 INICIAR SINCRONIZAÇÃO PTM26 → BDBI", type="primary", use_container_width=True, key="btn_sync"):
					with st.spinner("⏳ Sincronizando dados..."):
						# Criar backup antes
						try:
							backup_path = base_path / f'BD_backup_sync_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
							shutil.copy2(bd_path, backup_path)
							st.info(f"💾 Backup criado: {backup_path.name}")
						except Exception as e:
							st.warning(f"⚠️ Não foi possível criar backup: {str(e)}")
						
						# Executar sincronização
						sucesso, mensagem = sincronizar_ptm26_para_bdbi(bd_path)
						
						if sucesso:
							st.success(mensagem)
							st.balloons()
							
							st.info("🔄 Recarregando dados atualizados...")
							time.sleep(2)
							st.rerun()
						else:
							st.error(mensagem)
			
		except Exception as e:
			st.error(f"❌ Erro ao verificar abas: {str(e)}")
			st.warning("""
			⚠️ Certifique-se de que:
			1. O arquivo BD.xlsm existe
			2. Possui as abas 'BDBI' e 'PTM26'
			3. Ambas as abas têm as colunas 'Pedido' e 'Item'
			""")

	elif page == 'PTMs em atrasos':
		st.markdown('## 🔍 PTMs em atraso')
		st.markdown('Lista de PTMs com prazo vencido ou com entrega não registrada')
		st.markdown('---')

		df_remessas = carregar_dados_remessas(bd_path)

		if df_remessas.empty:
			st.warning("⚠️ Nenhum dado válido encontrado")
		else:
			df_remessas = calcular_metricas_remessas(df_remessas)
			# Seleciona atrasados (prazo vencido ou status VENCIDO)
			df_vencidos = df_remessas[df_remessas['Status'].str.contains('VENCIDO', na=False)]

			# Também considera sem entrega real e data prevista já passada
			mask_sem_entrega = (df_remessas['Entrega Real'].isna()) & (df_remessas['Dias até Entrega'].notna()) & (df_remessas['Dias até Entrega'] < 0)
			df_vencidos = pd.concat([df_vencidos, df_remessas[mask_sem_entrega]]).drop_duplicates()

			st.subheader(f"🔴 PTMs em atraso: {len(df_vencidos)}")

			if not df_vencidos.empty:
				df_show = df_vencidos.copy()
				# Calcula dias de atraso para ordenação
				if 'Dias até Entrega' in df_show.columns:
					df_show['Dias de Atraso'] = df_show['Dias até Entrega'].apply(lambda x: abs(int(x)) if pd.notna(x) else 0)
				else:
					df_show['Dias de Atraso'] = 0

				# Formata colunas
				for col in ['Data do Pedido', 'Data Prevista', 'Entrega Real']:
					if col in df_show.columns:
						df_show[col] = df_show[col].apply(formatar_data_abreviada)

				cols_exibir = [c for c in ['Pedido', 'Remessa', 'Dias de Atraso', 'Status', 'Data do Pedido', 'Data Prevista', 'Destino Ativo'] if c in df_show.columns]
				st.dataframe(df_show[cols_exibir].sort_values('Dias de Atraso', ascending=False).fillna(''), use_container_width=True, height=450)

				csv = df_show.to_csv(index=False, encoding='utf-8-sig')
				st.download_button(label="📥 Exportar PTMs em atraso (CSV)", data=csv, file_name=f"ptms_em_atraso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", mime='text/csv')
			else:
				st.success("✅ Nenhuma PTM em atraso no momento")

		st.markdown('---')

	# Rodapé personalizado
	st.markdown("---")
	st.markdown("<div style='text-align:center; font-size:11px; color:#b0c4de; margin-top:30px;'>🚀 Sistema PTM JSL 2.0 | Criado por Djalma A Barbosa (FYF9) | Todos os Direitos Reservados ® 2026</div>", unsafe_allow_html=True)


if __name__ == '__main__':
	import sys
	try:
		from streamlit.web import bootstrap
		if getattr(sys, 'frozen', False):
			script_path = sys.argv[0]
			bootstrap.run(script_path, "", [], {})
		else:
			main()
	except Exception:
		main()